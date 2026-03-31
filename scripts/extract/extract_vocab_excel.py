"""
해커스 노랭이 단어장 Excel에서 단어 데이터를 추출하여 JSON으로 변환한다.
입력: 00. Reference/hackers_vocab_basic.xlsx, hackers_vocab_800.xlsx, hackers_vocab_900.xlsx
출력: data/json/hackers_vocab.json (통합)
"""

import sys
import re
import json
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
REFERENCE_DIR = PROJECT_ROOT / "00. Reference"
OUTPUT_DIR = PROJECT_ROOT / "data" / "json"


def extract_기초(path: Path) -> list[dict]:
    """기초단어.xlsx: Sheet1 — [Day, 단어, 뜻, None, None]"""
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb["Sheet1"]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        day_str, word, meaning = row[0], row[1], row[2]
        if not word or not meaning:
            continue
        day_match = re.search(r"(\d+)", str(day_str))
        day = int(day_match.group(1)) if day_match else 0
        entries.append({
            "word": str(word).strip(),
            "meaning_kr": str(meaning).strip(),
            "day": day,
            "level": "기초",
        })
    wb.close()
    return entries


def extract_800(path: Path) -> list[dict]:
    """800점.xlsx: 30 sheets, two columns per sheet.
    Col[5]=num, Col[6]=word, Col[7]=meaning (group1)
    Col[14]=num, Col[15]=word, Col[16]=meaning (group2)
    """
    wb = openpyxl.load_workbook(str(path), read_only=True)
    entries = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        day_match = re.search(r"(\d+)", sheet_name)
        day = int(day_match.group(1)) if day_match else 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            # Group 1: columns 5,6,7
            if len(row) > 7 and row[6] and isinstance(row[6], str):
                word = row[6].strip()
                meaning = str(row[7]).strip() if row[7] else ""
                if word and meaning:
                    entries.append({
                        "word": word,
                        "meaning_kr": meaning,
                        "day": day,
                        "level": "800",
                    })
            # Group 2: columns 14,15,16
            if len(row) > 16 and row[15] and isinstance(row[15], str):
                word = row[15].strip()
                meaning = str(row[16]).strip() if row[16] else ""
                if word and meaning:
                    entries.append({
                        "word": word,
                        "meaning_kr": meaning,
                        "day": day,
                        "level": "800",
                    })
    wb.close()
    return entries


def extract_900(path: Path) -> list[dict]:
    """900점.xlsx: 30 sheets, single column.
    Col[5]=num, Col[6]=word, Col[7]=meaning
    Note: 'Day 25 (2)' sheet is actually Day 26.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True)
    entries = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 실제 Day 번호는 Row 0의 타이틀에서 추출
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        title = str(rows[0][0] or "")
        day_match = re.search(r"DAY\s*(\d+)", title, re.IGNORECASE)
        if day_match:
            day = int(day_match.group(1))
        else:
            day_match = re.search(r"(\d+)", sheet_name)
            day = int(day_match.group(1)) if day_match else 0

        for row in rows[1:]:
            row = list(row)
            if len(row) > 7 and row[6] and isinstance(row[6], str):
                word = row[6].strip()
                meaning = str(row[7]).strip() if row[7] else ""
                if word and meaning:
                    entries.append({
                        "word": word,
                        "meaning_kr": meaning,
                        "day": day,
                        "level": "900",
                    })
    wb.close()
    return entries


def assign_ids(entries: list[dict]) -> None:
    """hw_NNNN 형식 ID 부여"""
    for i, e in enumerate(entries, 1):
        e["id"] = f"hw_{i:04d}"


def save_json(data, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    print("=== 노랭이 단어장 Excel → JSON 변환 시작 ===\n")

    # 추출
    기초 = extract_기초(REFERENCE_DIR / "hackers_vocab_basic.xlsx")
    print(f"기초단어: {len(기초)}개")

    팔백 = extract_800(REFERENCE_DIR / "hackers_vocab_800.xlsx")
    print(f"800점: {len(팔백)}개")

    구백 = extract_900(REFERENCE_DIR / "hackers_vocab_900.xlsx")
    print(f"900점: {len(구백)}개")

    all_entries = 기초 + 팔백 + 구백
    assign_ids(all_entries)
    print(f"\n총 단어 수: {len(all_entries)}개")

    # 통합 파일 저장
    save_json(all_entries, OUTPUT_DIR / "hackers_vocab.json")
    print(f"저장: {OUTPUT_DIR / 'hackers_vocab.json'}")

    # 통계
    levels = defaultdict(int)
    for e in all_entries:
        levels[e["level"]] += 1
    print(f"\n레벨별 통계: {dict(levels)}")

    days = sorted(set(e["day"] for e in all_entries if e["day"] > 0))
    print(f"Day 범위: {min(days)} ~ {max(days)} ({len(days)}개 단원)")
    print("\n=== 변환 완료 ===")


if __name__ == "__main__":
    main()
